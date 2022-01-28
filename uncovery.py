import os
import sys
import requests
from requests.models import Response
from termcolor import cprint
from termcolor import colored
import colorama
import json
import pwinput
import numpy as np
from  openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import traceback
from datetime import date
import pickle
from mailPGP import Mail


#https://api.uncovery.io/v1/documentation

DEBUG = False
OBJ = ""
MSG = ""

def signin(email, password):
    cprint("Signin...","yellow")
    r = requests.post('https://api.uncovery.io/v1/auth/signin', data={'email': email,'password': password})
    response = r.json()
    #cprint(response,'cyan')
    if r.ok:
        cprint(response['request']['message'],"green")
        return response['data']['accessToken']
    else:
        cprint(response['message'],"red")
        sys.exit(0)

def createAndSendAuthRequest(token,url,payload):
    header = {'Authorization': 'Bearer ' + token}
    r = requests.get(url, params=payload,headers=header)
    response = r.json()
    #cprint(response,'cyan')
    return r,response
    
def getAllEntities(token,url,payload):
    cprint("Getting all Entities...","yellow")
    r,response = createAndSendAuthRequest(token,url,payload) 
    
    if r.ok:
        cprint(response['request']['message'],"green")
        entitiesDict = {}
        for entity in response['data']:
            print(colored(entity['name'], 'magenta') + colored(' : '+str(entity['id']),'yellow'))
            entitiesDict[entity['name']] = str(entity['id'])
        return entitiesDict
    else:
        cprint(response['message'],"red")
        sys.exit(0)

def getAllAssetsOfAnEntity(entitiesAndID,token,assetType):
    AssetsDict = {}
    for entity, id in entitiesAndID.items():
        pageNumber = 1
        ipsDict = {}    
        while True:
            print(colored("Getting all Assets of entity: ","yellow") + colored(entity, 'magenta') + colored(' Page : ','yellow') + colored(str(pageNumber),'cyan') + colored(' ...','cyan'))
            url = "https://api.uncovery.io/v1/entities/"+ str(id) +"/assets"
            payload = "type="+assetType+"&pageSize=20&page="+str(pageNumber)+"&sortBy=lastChanges&orderBy=desc"
            r,response = createAndSendAuthRequest(token,url,payload)
            if r.ok:
                cprint(response['request']['message'],"green")
           
                if str(response['pageInfo']['hasNextPage']) == 'False':
                    for item in response['data']:
                        ipsDict[item['id']] = item['value']
                    print(colored("No more page ","yellow") + colored(entity, 'magenta') + colored(' Page : ','yellow') + colored(str(pageNumber),'cyan') + colored(' !','yellow'))
                    break
                else:
                    for item in response['data']:
                        ipsDict[item['id']] = item['value']
                    pageNumber +=1
                    payload = "type="+(assetType)+"&pageSize=20&page="+str(pageNumber)+"&sortBy=lastChanges&orderBy=desc"
                    r,response = createAndSendAuthRequest(token,url,payload)
                    for item in response['data']:
                        ipsDict[item['id']] = item['value']
                    print(colored("Send request for page ","yellow") + colored(entity, 'magenta') + colored(' Page : ','yellow') + colored(str(pageNumber),'cyan'))
            else:
                cprint(response['message'],"red")
                break
        AssetsDict[id] = ipsDict
    return AssetsDict

def getOneAssetGraph(entitiesAndID,token,assetsAndEntitiesID):
    result = {}
    for entity, id in entitiesAndID.items():
        print(colored("For Entity : ","yellow") + colored(entity, 'magenta') + colored(' ...','yellow'))    
        entityDict = {}
        for idEntity, ipsDict in assetsAndEntitiesID.items():
            if id == idEntity:
                for ipId, ipValue in ipsDict.items():
                    portTcpArray = []
                    portUdpArray = []
                    portsDict = {}
                    print(colored("Getting Graph of asset: ","yellow") + colored(ipValue, 'cyan') + colored(' ...','yellow'))
                    url = "https://api.uncovery.io/v1/entities/"+ str(id) +"/assets/" + str(ipId) + "/graph"
                    payload = "direction=out"
                    r,response = createAndSendAuthRequest(token,url,payload)
                    if r.ok:
                        cprint(response['request']['message'],"green")

                        for item in response['data']['nodes']:
                            tmpDict = {}
                            tmpDict['port'] = item['value']
                            tmpDict['detail'] = "FIXME"
                            if item['type'] == 'porttcp':
                                portTcpArray.append(tmpDict)
                            if item['type'] == 'portudp':
                                portUdpArray.append(tmpDict)
                        portsDict['TCP'] = portTcpArray
                        portsDict['UDP'] = portUdpArray
                        entityDict[ipValue] = portsDict
                    else:
                        cprint(response['message'],"red")
                        break
        result[entity] = entityDict
    return result

def getDifferentsPorts(jsonFile):
    result = {}
    cprint('Retrieve every used ports...','yellow')
    try:
        with open(jsonFile,'r') as f_in:
            data = json.load(f_in)
            for entity in data:
                print(colored('For entity ','yellow') + colored(entity, 'magenta') + colored(' :','yellow'))
                TCPportArray = []
                UDPportArray = []
                portsDict = {}   
                for ip in data[entity]:
        
                    for tcp in data[entity][ip]["TCP"]:
                        TCPportArray.append(int(tcp['port']))
                    for udp in data[entity][ip]["UDP"]:
                        UDPportArray.append(int(udp['port']))
                TCPraw = np.array(TCPportArray)
                UDPraw = np.array(UDPportArray)
                TCPuniquePorts = np.unique(TCPraw)
                UDPuniquePorts = np.unique(UDPraw)
                TCPsortedAndUniquePorts = sorted(TCPuniquePorts)
                UDPsortedAndUniquePorts = sorted(UDPuniquePorts)
                portsDict["TCP"] = TCPsortedAndUniquePorts
                portsDict["UDP"] = UDPsortedAndUniquePorts
                result[entity] = portsDict
                print(colored('TCP: ','magenta') + colored(TCPsortedAndUniquePorts,'cyan'))
                print(colored('UDP: ','magenta') + colored(UDPsortedAndUniquePorts,'cyan'))
        print(colored('Success retrieving ports','green'))
    except:
        print(colored('Error retrieving ports','red'))
        print(traceback.format_exc() if DEBUG else '')
    return result

def cleanSubdirectory(directory):
    print(colored("Removing files in \" ",'yellow') + colored(directory,'magenta') + colored(" \" directory...","yellow"))
    if os.path.exists(directory):
        for files in os.listdir(directory):
            os.remove(os.path.join(directory, files))
    else:
        try:
            os.mkdir(directory)
            print(colored('Cleaning success','green'))
        except Exception:
            print(colored("Error : could not create " + directory + " subdirectory !",'red'))

def createExcelSheets(entitiesAndID,portList,jsonFile):
    cleanSubdirectory('output')
    wb1 = Workbook()
    try:
        print(colored('Creating Excel output...','yellow'))
        for entity,id in entitiesAndID.items():
            actualPortIndex = actualIpIndex = 2
            ws = wb1.create_sheet(entity)
            with open(jsonFile,'r') as f_in:
                data = json.load(f_in)
                for ip in data[entity]:
                    ws.cell(row = actualIpIndex, column = 1).value = str(ip)
                    for protocol in portList[entity]:
                        for port in portList[entity][protocol]:
                            ws.cell(row = 1, column = actualPortIndex).value = protocol.lower()+'/'+str(port)
                            ws.cell(row = 1, column = actualPortIndex).alignment = Alignment(horizontal='center', textRotation = 90)
                            ws.column_dimensions[get_column_letter(actualPortIndex)].width = 3
                            #if (str(port) in data[entity][ip][protocol]): # A FIX, condition de detection
                            if any(x['port'] == str(port) for x in data[entity][ip][protocol]):    
                                ws.cell(row = actualIpIndex, column = actualPortIndex).value = 'X'
                                ws.cell(row = actualIpIndex, column = actualPortIndex).alignment = Alignment(horizontal='center')
                            actualPortIndex = actualPortIndex + 1
                    actualPortIndex = 2 # reset Index port
                    actualIpIndex = actualIpIndex + 1
            ws.row_dimensions[1].height = 50
            ws.freeze_panes = "A2"
            ws.column_dimensions['A'].width = 15
        wb1.save(os.path.join('output','output.xlsx'))
        print(colored('Success when creating Excel file','green'))
    except:
        print(colored('Error when creating Excel file','red'))
        print(traceback.format_exc() if DEBUG else '')

def diff(obj1, obj2):
    change = []
    today = date.today().strftime("%Y/%m/%d")

    try:
        with open('changes-history.picle', 'rb') as f:
            changes_history = pickle.load(f)
    except:
        # fichier non existant, sera cree a la fin du diff
            changes_history = [{'type':'None', 'ip':'None', 'port':'None', 'protocol':'None', 'timestamp':today}]
            print(traceback.format_exc() if DEBUG else '')
    
    # vérifier si des ips ont ete ajoutees ou supprimes dans le scope

    del_ip = [ x for x in obj1 if not x in obj2 ]
    add_ip = [ x for x in obj2 if not x in obj1 ]

    for i in del_ip:
        change.append({'type':'removeip', 'ip':i, 'timestamp':today})

    for i in add_ip:
        change.append({'type':'addip', 'ip':i, 'timestamp':today})



    #  verifier les differences sur les ports
    interlist = obj1.keys() & obj2.keys()

    for protocol in ['TCP', 'UDP']:
        for i in interlist :
            ports1 = [ x.get('port') for x in obj1.get(i).get(protocol)]
            ports2 = [ x.get('port') for x in obj2.get(i).get(protocol)]

            add_ports = [ x for x in ports2 if not x in ports1]
            del_ports = [ x for x in ports1 if not x in ports2]

            for p in add_ports:
                change.append({'type':'addport', 'ip':i, 'port':p, 'protocol':protocol, 'timestamp':today})

            for p in del_ports:
                change.append({'type':'removeport', 'ip':i, 'port':p, 'protocol':protocol, 'timestamp':today})

    all_changes = [*change, *changes_history]
    with open('changes-history.picle', 'wb') as f:
        pickle.dump(all_changes,f)


    global OBJ
    global MSG
    if(len(all_changes) > len(changes_history)): # si nouveau changement ajouté
        MSG = genMessage(all_changes,len(changes_history))
        OBJ = "Uncovery Daily Update - Changements detectes !"
    else:
        MSG = "Pas de nouveaux changements détectés :)"
        OBJ = "Uncovery Daily Update - Rien a signaler !"
    return all_changes

def getDiffBetweenEntity(entitiesAndID,previous,actual):

    previousData = actualData = None
    with open(previous,'r') as previousFile:
            previousData = json.load(previousFile)
    with open(actual,'r') as actualFile:
            actualData = json.load(actualFile)
    
    obj1 = {}
    obj2 = {}

    for i in previousData.keys():
        obj1.update(previousData[i])

    for i in actualData.keys():
        obj2.update(actualData[i])

    changes = diff(obj1,obj2)
        
    for x in changes:
        if 'port' in x['type']:
            sendUpdateToExcel(x['timestamp'],x['type'],x['ip'],x['protocol'],x['port'])
        else:
            sendUpdateToExcel(x['timestamp'],x['type'],x['ip'])

        # osef, juste pour prompt joli
        if 'remove' in x['type']:
            if 'ip' in x['type']:
                print(colored("Delete: ","red")+colored("ip: ","yellow")+colored(x['ip'],"cyan"))
            else:
                print(colored("Delete: ","red")+colored("port: ","yellow")+colored(x['protocol']+"/","cyan")+colored(x['port'],"cyan")+colored(" ip: ","yellow")+colored(x['ip'],"cyan"))
        else:
            if 'ip' in x['type']:
                print(colored("Add: ","green")+colored("ip: ","yellow")+colored(x['ip'],"cyan"))
            else:
                print(colored("Add: ","green")+colored("port: ","yellow")+colored(x['protocol']+"/","cyan")+colored(x['port'],"cyan")+colored(" ip: ","yellow")+colored(x['ip'],"cyan"))
    
def sendUpdateToExcel(*args):
    wb = load_workbook(os.path.join('output','output.xlsx'))
    wb.worksheets[0].title = 'Historique des changements'
    ws = wb.worksheets[0]
    ws.insert_rows(1)
    
    actualCol = 1
    for x in args:
        ws.cell(row = 1, column = actualCol).value = x
        ws.cell(row = 1, column = actualCol).alignment = Alignment(horizontal='center')
        actualCol = actualCol +1
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 7
    wb.save(os.path.join('output','output.xlsx'))                   


def genMessage(all_changes,indice):
    msg = "Les changements sont: \n"
    for x in all_changes[:-indice]:
        # juste pour mail joli
        if 'remove' in x['type']:
            if 'ip' in x['type']:
                msg += "Delete ip: "+x['ip']+"\n"
            else:
                msg += "Delete port: "+x['protocol']+"/"+x['port']+" from ip: "+x['ip']+"\n"
        else:
            if 'ip' in x['type']:
                msg += "Add ip: "+x['ip']+"\n"
            else:
                msg += "Add port: "+x['protocol']+"/"+x['port']+" from ip: "+x['ip']+"\n"
    return msg

if __name__ == '__main__':
    colorama.init()
    # check arguments
    email = password = ''
    if ((len(sys.argv) == 2) or (len(sys.argv) > 3)):
        cprint("Error, usage is: python {0} <email> <password>".format(sys.argv[0]),"red")
        cprint("Error, usage is: python {0} ".format(sys.argv[0]),"red")
        sys.exit(1)
    elif (len(sys.argv) == 3):
        email = sys.argv[1]
        password = sys.argv[2]
    else:
        email = input("Email: ")
        password = pwinput.pwinput(prompt='Password: ', mask='*')
    
    accessToken = signin(email,password)
    entitiesAndID = getAllEntities(accessToken,url = 'https://api.uncovery.io/v1/entities', payload = {'pageSize':'100','sortBy':'name','orderBy':'asc'})
    assetsAndEntitiesID = getAllAssetsOfAnEntity(entitiesAndID,accessToken,"ipv4")
    openedPortsByIPForEachEntity = getOneAssetGraph(entitiesAndID,accessToken,assetsAndEntitiesID)
    jsonData = json.dumps(openedPortsByIPForEachEntity, sort_keys=True)
    orig_stdout = sys.stdout
    sys.stdout = open('data.json', 'w+')
    print(jsonData) # Save every open tcp/udp ports
    sys.stdout = orig_stdout
    portListForEachEntity = getDifferentsPorts('data.json') # list of tcp/udp used ports
    createExcelSheets(entitiesAndID,portListForEachEntity,'data.json')
        
    try:
        print(colored('Compute differences history...','yellow'))
        getDiffBetweenEntity(entitiesAndID,"previous.json","data.json")
        try:
            print(colored("Sending e-mail ...","yellow"))
            print(colored("=================================","magenta"))
            print(colored("OBJ: ","magenta")+colored(OBJ,"yellow"))
            print(colored("CONTENT: ","magenta")+colored(MSG,"yellow"))
            print(colored("=================================","magenta"))
            m = Mail(OBJ,MSG,os.path.join('output','output.xlsx'))
            m.sendmail()
            print(colored("Mail has been send !","green"))
        except:
            print(colored("Error when sending email","red"))
        os.remove('previous.json')
    except:
        print(colored('Warning: no previous run !','red'))
        print(traceback.format_exc() if DEBUG else '')
    os.rename('data.json', 'previous.json')
    print(colored('Finish !','green'))
